from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service()
serv.creation_flags=CREATE_NO_WINDOW
driver=webdriver.Chrome(service=serv)



import time, datetime, os
import openpyxl   


now = str(datetime.datetime.now())[:16]
folderName = now.replace(":", "_")
os.mkdir(folderName)


key_word = ["공모주", "공모주 일정"]

wb = openpyxl.Workbook()  #워크북 열기

for i in range(len(key_word)):
    ws = wb.create_sheet()   #워크시트 만들기
    ws.title = key_word[i]
    ws.column_dimensions["A"].width = 90   #A열 조절
    
    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i]
    driver.get(url)  #검색어 url 검색
    time.sleep(2)

    list_news = driver.find_element("class name", "list_news")
    news_boxes = list_news.find_elements("class name", "bx")
    
    for j in range(len(news_boxes)):      
        driver.execute_script("arguments[0].scrollIntoView(true);", news_boxes[j])   #캡처할곳으로 스크롤
        file = f"{folderName}/{i+1}_{key_word[i]}-{j+1}.png"  #파일 이름 지정
        news_boxes[j].screenshot(file)   #스크린샷 이미지 저장

        ws.row_dimensions[j+1].height = 100   #행 높이 조절하기
        img = openpyxl.drawing.image.Image(file)   #액셀에 이미지 삽입 준비
        ws.add_image(img, f'A{j+1}')   #A열에 이미지  넣기

        title = news_boxes[j].find_element("class name", "news_tit")   #엑셀 파일에 링크 삽입 할 준비
        print(j+1, title.text)
        link = title.get_attribute("href")   #title 개체의 링크를 저장
        ws[f'B{j+1}'].value = "기사링크"   
        ws[f'B{j+1}'].hyperlink = link   #B열 해당 셀(j+1)에 링크 삽입


    print()
    
wb.remove(wb["Sheet"])   #필요없는 첫번째 시트 삭제
wb.save(f"{folderName}/{folderName}_결과.xlsx")   #엑셀 저장

        


